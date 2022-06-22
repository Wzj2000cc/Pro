import json
import os
import base64
import pyqrcode
import cv2

from glob import glob
from pathlib import Path
from svglib.svglib import svg2rlg
from reportlab.graphics import renderPM

from flask import Blueprint, request, Response
from common.utils import common_method, ck_login, validate
from common.utils.logs import logger
from applications.models.auth_model import db, User, Code_two

AI_blu = Blueprint('ai', __name__)

dst_dir = os.getcwd() + '/applications/view/AI/'
file_dir = os.getcwd() + '/applications/file'
role_dict = {"admin": "管理员", "user": "审计用户", "person": "部门专责"}


# 文件加密函数
@AI_blu.route('/lockfile/', methods=['GET'])
def Lock_file(key_path=None, *, encoding='utf-8'):
    uname = validate.xss_escape(request.args.get('uname'))
    file_name = validate.xss_escape(request.args.get('file_name'))

    # ======== 安全登录校验 ========
    if ck_login.is_status(uname) is None:
        return Response('该用户不存在，请注册或换账号登录！')
    if not ck_login.is_status(uname):
        return Response('当前状态为离线，请重新登录！')
    if not uname:
        return Response('用户名输入为空，请重新输入')
    # ================

    path = Path(dst_dir + file_name)
    cwd = path.cwd() / 'applications' / 'view' / 'AI' / file_name.split('.')[0]
    path_encrypted = cwd / path.name

    if key_path is None:
        key_path = cwd / 'key'
    if not cwd.exists():
        cwd.mkdir()
        path_encrypted.touch()
        key_path.touch()

    # 打开文件并将内容写入
    try:
        with path.open('rt', encoding=encoding) as f1, \
                path_encrypted.open('wt', encoding=encoding) as f2, \
                key_path.open('wt', encoding=encoding) as f3:
            encrypted, key = common_method.encrypt(f1.read())
            json.dump(encrypted, f2)
            json.dump(key, f3)
    except Exception as e:
        logger.error(e)
        return Response('加密文件不支持上传类型')

    # 将加密后的文件进行覆盖
    src_dir = str(cwd) + '/'
    src_file_list = glob(src_dir + '*')  # glob获得路径下所有文件，可根据需要修改
    for srcfile in src_file_list:
        common_method.mymovefile(srcfile, dst_dir)  # 移动文件
    os.rmdir(cwd)

    logger.info(f'用户（{uname}）对文件{path.name}加密成功')
    return Response(f'用户（{uname}）对文件{path.name}加密成功')


@AI_blu.route('/unlockfile/', methods=['GET'])
# 文件解密函数
def decrypt_file(key_path=None, *, encoding='utf-8'):
    uname = validate.xss_escape(request.args.get('uname'))
    file_name = validate.xss_escape(request.args.get('file_name'))

    # ======== 安全登录校验 ========
    if ck_login.is_status(uname) is None:
        return Response('该用户不存在，请注册或换账号登录！')
    if not ck_login.is_status(uname):
        return Response('当前状态为离线，请重新登录！')
    if not uname:
        return Response('用户名输入为空，请重新输入')
    # ================

    path_encrypted = Path(dst_dir + file_name)
    cwd = path_encrypted.cwd() / 'applications' / 'view' / 'AI'
    path_decrypted = cwd / 'decrypted'

    if not path_decrypted.exists():
        path_decrypted.mkdir()
        path_decrypted /= path_encrypted.name
        path_decrypted.touch()
    if key_path is None:
        key_path = cwd / 'key'

    try:
        with path_encrypted.open('rt', encoding=encoding) as f1, \
                key_path.open('rt', encoding=encoding) as f2, \
                path_decrypted.open('wt', encoding=encoding) as f3:
            decrypted = common_method.decrypt(json.load(f1), json.load(f2))
            f3.write(decrypted)
    except Exception as e:
        logger.error(e)
        return Response('该文件未进行加密或不支持上传文件的类型')

    # 将解密后的文件进行覆盖
    src_dir = dst_dir + 'decrypted/'
    src_file_list = glob(src_dir + '*')  # glob获得路径下所有文件，可根据需要修改
    for srcfile in src_file_list:
        common_method.mymovefile(srcfile, dst_dir)  # 移动文件
    src_dir = dst_dir + 'decrypted'
    os.rmdir(src_dir)

    logger.info(f'用户（{uname}）对加密文件{path_encrypted.name}解密成功')
    return Response(f'用户（{uname}）对加密文件{path_encrypted.name}解密成功')


@AI_blu.route('/code_two/', methods=['POST'])
def Code_Two():
    data = request.get_data(as_text=True)
    json_data = json.loads(data)
    uname = validate.xss_escape(json_data.get('uname'))
    to_path = validate.xss_escape(json_data.get('to_path'))
    code_name = validate.xss_escape(json_data.get('code_name'))

    # ======== 安全登录校验 ========
    if ck_login.is_status(uname) is None:
        return Response('该用户不存在，请注册或换账号登录！')
    if not ck_login.is_status(uname):
        return Response('当前状态为离线，请重新登录！')
    if not uname:
        return Response('用户名输入为空，请重新输入')
    # ================

    userid = None
    role_code = None
    users = User.query.filter_by(uname=uname).all()
    for user in users:
        userid = user.id
        role_code = role_dict.get(user.role)
    codes = Code_two.query.filter_by(id=userid).all()
    for code in codes:
        if code_name + '.svg' == code.code_name:
            return Response(f'二维码名称（{code_name}）已存在，请更换后重试')
        role_code = role_dict.get(code.user.role)

    # 生成svg二维码
    url = pyqrcode.create(to_path)
    url.svg(code_name + '.svg', scale=8)

    # svg转png
    cwd = os.getcwd() + '/' + code_name + '.svg'
    drawing = svg2rlg(cwd)
    renderPM.drawToFile(drawing, code_name + '.png', fmt="PNG")

    # 图片转换base64格式返回template
    img_local_path = os.getcwd() + '/' + code_name + '.png'
    with open(img_local_path, 'rb') as img_f:
        img_stream = img_f.read()
        img_stream = base64.b64encode(img_stream)

    code_two = Code_two(code_name=code_name + '.svg', to_path=to_path, svg_path=cwd,
                        png_base64='data:image/png;base64,' + img_local_path, usercode=userid)

    # 移动文件至file
    common_method.mymovefile(cwd, file_dir)
    common_method.mymovefile(img_local_path, file_dir)

    try:
        db.session.add(code_two)
        db.session.commit()
        logger.info(f"{role_code}（{uname}） 生成二维码（{code_name}）成功!")
        return Response(f'{role_code}（{uname}） 生成二维码（{code_name}）成功! \n'
                        'data:image/png;base64,' + img_stream.decode('utf-8'))
    except Exception as e:
        logger.info(f"{role_code}（{uname}） 生成二维码（{code_name}）失败!")
        logger.error(e)
        return Response(f'{role_code}（{uname}） 生成二维码（{code_name}）失败!')


# img转换素描图
@AI_blu.route('/wmi/', methods=['GET'])
def System_spec():
    img_name = validate.xss_escape(request.args.get('img_name'))
    # 读取图片
    img = cv2.imread(file_dir + "/" + img_name + ".jpg")
    # 灰度
    grey = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    invert = cv2.bitwise_not(grey)
    # 高斯滤波
    blur_img = cv2.GaussianBlur(invert, (7, 7), 0)
    inverse_blur = cv2.bitwise_not(blur_img)
    sketch_img = cv2.divide(grey, inverse_blur, scale=256.0)
    # 保存
    cv2.imwrite(img_name + '1.jpg', sketch_img)
    cv2.waitKey(0)
    cv2.destroyAllWindows()
    # 移动文件至file
    common_method.mymovefile(os.getcwd() + '/' + img_name + '1.jpg', file_dir + '/')

    logger.info(f'素描图（{img_name}）生成成功')
    return Response('素描图生成成功')


@AI_blu.route('/test/', methods=['GET'])
def Test():
    from PIL import Image
    import openpyxl
    from openpyxl.styles import PatternFill, Fill

    imageFileName = ''  # 图片文件名
    image = Image.open(imageFileName)  # 打开图片
    wb = openpyxl.Workbook()  # 创建Excel
    sheet = wb.create_sheet(imageFileName)  # 创建sheet
    imgW, imgH = image.size  # 获取图片大小
    for w in range(imgW):
        for h in range(imgH):
            # 将每个像素的颜色填充到对应cell的背景色中
            rgba = image.getpixel((w, h))
            colorHex = hex(rgba[0])[2:].zfill(2) + hex(rgba[1])[2:].zfill(2) + hex(rgba[2])[2:].zfill(2)
            fill = PatternFill(fill_type='solid', start_color=colorHex, end_color=colorHex)
            sheet.cell(row=h + 1, column=w + 1).fill = fill
    wb.save(imageFileName + '.xlsx')  # 保存xlsx文件